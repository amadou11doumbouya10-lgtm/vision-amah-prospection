"""
Prospection Vision Amah — commerces sans site web à Conakry
===========================================================

Un seul programme : il cherche sur OpenStreetMap, nettoie les données,
déduplique, rédige un message par prospect, puis OUVRE AUTOMATIQUEMENT le
résultat en PDF. Rien d'autre à installer, aucune clé API, aucune carte
bancaire.

Il NE FAIT AUCUN ENVOI. L'envoi reste manuel : automatiser WhatsApp expose le
numéro à un bannissement, et c'est justement le numéro affiché sur le site.

    pip install requests fpdf2 openpyxl
    python prospect_finder.py

Le PDF (prospects.pdf) est le format ouvert automatiquement : contrairement à
une page web, il s'ouvre de façon fiable avec le lecteur PDF par défaut, sans
dépendre du navigateur. Un fichier prospects.csv (classement/sauvegarde) et
une page prospects.html (interface interactive : recherche, filtres, statuts,
relances, stats, pages de démo) sont produits en plus, à ouvrir à la main si
besoin.

Le bouton « Exporter le journal » de prospects.html enregistre un CSV listant
chaque prospect dont le statut a changé (donc a priori déjà contacté), avec
ses informations et le message envoyé. Ce script l'absorbe automatiquement au
lancement suivant dans un classeur Excel dédié (« Journal des envois - Vision
Amah.xlsx »), géré entièrement par ce programme — jamais dans un fichier de
suivi manuel existant, pour ne prendre aucun risque avec des notes ou mises en
forme déjà en place.
"""

import csv
import html
import json
import math
import os
import pathlib
import re
import sys
import time
import unicodedata
import urllib.parse
import webbrowser

import requests

# La console Windows utilise par défaut l'encodage cp1252 pour la sortie
# standard, qui ne connaît ni les emojis des libellés ni certains signes
# typographiques (ex. le tiret « − ») utilisés dans les messages du
# programme : sans ceci, un print() sur l'un de ces caractères fait planter
# le script avec une UnicodeEncodeError, y compris après avoir déjà récupéré
# les résultats d'OpenStreetMap.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# overpass-api.de sature régulièrement (trafic de bots IA, entre autres) et
# répond alors 504. kumi.systems est un miroir public gratuit, sans inscription,
# maintenu spécifiquement pour absorber ce genre de surcharge — on l'utilise en
# secours si l'instance principale échoue.
OVERPASS_ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
]
FICHIER_CSV = "prospects.csv"
FICHIER_HTML = "prospects.html"
FICHIER_PDF = "prospects.pdf"
FICHIER_JOURNAL_EXPORT = "journal_envois.csv"
FICHIER_JOURNAL_XLSX = "Journal des envois - Vision Amah.xlsx"

# Le suivi « déjà contactés » doit survivre d'une exécution à l'autre — il ne
# peut donc pas dépendre du dossier où se trouve le script (souvent le Bureau,
# régulièrement bloqué en écriture par le Contrôle d'accès aux dossiers de
# Windows Defender). On le range dans un dossier applicatif toujours
# accessible, à un chemin fixe.
DOSSIER_APP = os.path.join(
    os.environ.get("LOCALAPPDATA") or os.path.expanduser("~"), "ProspectionVisionAmah"
)
os.makedirs(DOSSIER_APP, exist_ok=True)
FICHIER_DEJA_VUS = os.path.join(DOSSIER_APP, "deja_contactes.json")

SIGNATURE = "Amadou — Vision Amah — +224 611 76 85 52"
SITE = "vision-amah.vercel.app"

# ---------------------------------------------------------------- catégories

FILTRES_OSM = [
    '"amenity"~"^(restaurant|cafe|fast_food|bar|pharmacy|fuel|bank|car_wash|car_rental|dentist|doctors|clinic|veterinary|driving_school|language_school|nightclub|cinema|events_venue|conference_centre)$"',
    '"shop"',
    '"tourism"~"^(hotel|guest_house|hostel|motel)$"',
    '"office"',
    '"craft"',
    '"leisure"~"^(fitness_centre|sports_centre|water_park)$"',
]

FAMILLES = {
    "restauration": {"restaurant", "cafe", "fast_food", "bar", "bakery", "pastry",
                     "butcher", "greengrocer", "confectionery"},
    "beaute": {"beauty", "hairdresser", "cosmetics", "perfumery"},
    "mode": {"clothes", "boutique", "tailor", "shoes", "jewelry",
             "fashion_accessories", "bag"},
    "informatique": {"computer", "mobile_phone", "electronics", "electrical"},
    "alimentation": {"supermarket", "convenience", "mall", "wholesale"},
    "sante": {"pharmacy", "dentist", "doctors", "clinic", "veterinary"},
    "hebergement": {"hotel", "guest_house", "hostel", "motel"},
    "auto": {"fuel", "car_wash", "car_rental", "car_repair"},
    "services": {"driving_school", "language_school", "travel_agency", "travel_agent",
                 "estate_agent", "lawyer", "accountant", "insurance", "employment_agency"},
    "loisirs": {"nightclub", "cinema", "events_venue", "conference_centre",
                "fitness_centre", "sports_centre", "water_park"},
}

BESOIN = {
    "restauration": "automatisation",
    "beaute":       "automatisation",
    "sante":        "automatisation",
    "hebergement":  "automatisation",
    "auto":         "automatisation",
    "services":     "automatisation",
    "mode":         "site",
    "alimentation": "site",
    "informatique": "site",
    "loisirs":      "site",
    "generique":    "site",
}

LABELS_GROUPE = {
    "restauration": "🍽️ Restauration",
    "beaute": "💇 Beauté",
    "mode": "👗 Mode",
    "alimentation": "🛒 Alimentation",
    "informatique": "💻 Informatique",
    "sante": "🏥 Santé",
    "hebergement": "🏨 Hébergement",
    "auto": "🚗 Automobile",
    "services": "💼 Services professionnels",
    "loisirs": "🎉 Loisirs",
    "generique": "🏪 Autres services",
}
ORDRE_AUTO    = ["restauration", "beaute", "sante", "hebergement", "auto", "services"]
ORDRE_SITE    = ["mode", "alimentation", "informatique", "loisirs", "generique"]
ORDRE_GROUPES = ORDRE_AUTO + ORDRE_SITE


def famille(valeur_osm):
    for nom, valeurs in FAMILLES.items():
        if valeur_osm in valeurs:
            return nom
    return "generique"


# ------------------------------------------------------------------ messages

# Chaque texte annonce explicitement l'aperçu joint (la démo générée par
# ecrire_demos/corps_demo) et décrit EXACTEMENT ce qu'elle montre pour cette
# famille — menu chiffré, catalogue photo, liste de prestations réservables,
# ou agenda — pour que le message et la pièce jointe racontent la même chose
# plutôt que deux pitchs génériques qui ne se recoupent pas forcément.
MESSAGES = {
    "restauration": (
        "Bonjour {nom} 👋\n\n"
        "Vos commandes arrivent encore par téléphone, une par une. Je vous ai "
        "préparé un aperçu (en pièce jointe) : votre menu en ligne avec commande "
        "directe par WhatsApp, confirmation automatique envoyée au client, et "
        "suivi de commande sans appel. Vos clients commandent sans vous déranger, "
        "vous gagnez du temps sur chaque service.\n\n"
        "Mes réalisations : {site}\n\n"
        "Si ça vous intéresse, je vous explique en deux minutes.\n\n{signature}"
    ),
    "beaute": (
        "Bonjour {nom} 👋\n\n"
        "Chaque prise de rendez-vous vous prend un appel et une saisie manuelle. "
        "Je vous ai préparé un aperçu (en pièce jointe) : vos prestations en "
        "ligne, réservation directe par WhatsApp, et rappel automatique envoyé "
        "à votre cliente la veille. Moins d'absences, moins de téléphone, plus "
        "de temps pour travailler.\n\n"
        "Mes réalisations : {site}\n\n"
        "Si ça vous intéresse, je vous explique en deux minutes.\n\n{signature}"
    ),
    "mode": (
        "Bonjour {nom} 👋\n\n"
        "Vos articles ne sont visibles que par ceux qui passent devant la "
        "boutique. Je vous ai préparé un aperçu (en pièce jointe) : votre "
        "catalogue avec photos et prix, et une commande directe par WhatsApp.\n\n"
        "Mes réalisations : {site}\n\n"
        "Si ça vous intéresse, je vous explique en deux minutes.\n\n{signature}"
    ),
    "informatique": (
        "Bonjour {nom} 👋\n\n"
        "Vous vendez de la technologie, mais vos clients ne trouvent ni vos "
        "produits ni vos tarifs en ligne. Je vous ai préparé un aperçu (en pièce "
        "jointe) : votre catalogue produits et une demande de devis directe.\n\n"
        "Mes réalisations : {site}\n\n"
        "Si ça vous intéresse, je vous explique en deux minutes.\n\n{signature}"
    ),
    "alimentation": (
        "Bonjour {nom} 👋\n\n"
        "Vos clients ne savent pas ce que vous avez en rayon avant de se "
        "déplacer. Je vous ai préparé un aperçu (en pièce jointe) : vos produits "
        "en photo, vos promotions, et un contact direct.\n\n"
        "Mes réalisations : {site}\n\n"
        "Si ça vous intéresse, je vous explique en deux minutes.\n\n{signature}"
    ),
    "sante": (
        "Bonjour {nom} 👋\n\n"
        "Vos patients appellent pour savoir si vous êtes disponibles, et oublient "
        "souvent leur rendez-vous. Je vous ai préparé un aperçu (en pièce jointe) "
        ": prise de rendez-vous en ligne et rappel automatique envoyé la veille "
        "par WhatsApp. Moins d'absences, moins d'appels à gérer, plus de "
        "consultations honorées.\n\n"
        "Mes réalisations : {site}\n\n"
        "Si ça vous intéresse, je vous explique en deux minutes.\n\n{signature}"
    ),
    "hebergement": (
        "Bonjour {nom} 👋\n\n"
        "Vos réservations passent encore par appel téléphonique, sans confirmation "
        "automatique. Je vous ai préparé un aperçu (en pièce jointe) : réservation "
        "en ligne, confirmation instantanée par WhatsApp, et rappel automatique "
        "la veille de l'arrivée. Moins de no-shows, moins d'appels, plus de "
        "chambres remplies.\n\n"
        "Mes réalisations : {site}\n\n"
        "Si ça vous intéresse, je vous explique en deux minutes.\n\n{signature}"
    ),
    "auto": (
        "Bonjour {nom} 👋\n\n"
        "Vos clients rappellent pour savoir si leur véhicule est prêt, et les "
        "rendez-vous se perdent dans des appels en double. Je vous ai préparé un "
        "aperçu (en pièce jointe) : prise de rendez-vous en ligne, confirmation "
        "automatique, et message WhatsApp envoyé quand le véhicule est prêt. "
        "Moins d'appels, plus de clients fidèles.\n\n"
        "Mes réalisations : {site}\n\n"
        "Si ça vous intéresse, je vous explique en deux minutes.\n\n{signature}"
    ),
    "services": (
        "Bonjour {nom} 👋\n\n"
        "Vos relances clients et vos suivis de dossiers se font encore à la main, "
        "un par un. Je vous ai préparé un aperçu (en pièce jointe) : présentation "
        "de vos services en ligne, formulaire de contact direct, et relances "
        "automatiques par WhatsApp. Vous vous concentrez sur le travail, les "
        "suivis se font tout seuls.\n\n"
        "Mes réalisations : {site}\n\n"
        "Si ça vous intéresse, je vous explique en deux minutes.\n\n{signature}"
    ),
    "loisirs": (
        "Bonjour {nom} 👋\n\n"
        "Vos clients ne trouvent pas votre programme ni vos horaires en ligne "
        "avant de venir. Je vous ai préparé un aperçu (en pièce jointe) : votre "
        "programme toujours à jour, et un contact direct par WhatsApp.\n\n"
        "Mes réalisations : {site}\n\n"
        "Si ça vous intéresse, je vous explique en deux minutes.\n\n{signature}"
    ),
    "generique": (
        "Bonjour {nom} 👋\n\n"
        "Je n'ai pas trouvé de site web pour votre activité — vos clients ne "
        "peuvent donc pas vous découvrir en ligne. Je vous ai préparé un aperçu "
        "(en pièce jointe) : la présentation de votre activité et un contact "
        "facilité.\n\n"
        "Mes réalisations : {site}\n\n"
        "Si ça vous intéresse, je vous explique en deux minutes.\n\n{signature}"
    ),
}

# ------------------------------------------------------------- pages de démo

COULEUR_GROUPE = {
    "restauration": "#f97316", "beaute": "#ec4899", "mode": "#a855f7",
    "informatique": "#3b82f6", "alimentation": "#22c55e", "sante": "#14b8a6",
    "hebergement": "#eab308", "auto": "#64748b", "services": "#6366f1",
    "loisirs": "#f43f5e", "generique": "#3b82f6",
}

ICONES_GROUPE = {
    "restauration": "🍽️", "beaute": "💇", "mode": "👗", "informatique": "💻",
    "alimentation": "🛒", "sante": "🏥", "hebergement": "🏨", "auto": "🚗",
    "services": "💼", "loisirs": "🎉", "generique": "🏪",
}

TAGLINES = {
    "restauration": "Découvrez notre menu et commandez en toute simplicité.",
    "beaute": "Réservez votre prestation en quelques clics.",
    "mode": "Notre collection, toujours à portée de main.",
    "informatique": "Nos produits et nos prix, sans vous déplacer.",
    "alimentation": "Ce qui vous attend en rayon, avant même de venir.",
    "sante": "Nos horaires et nos coordonnées, disponibles à tout moment.",
    "hebergement": "Vos prochaines nuits commencent ici.",
    "auto": "Vos prestations auto, sans prise de tête.",
    "services": "Un service de confiance, à portée de clic.",
    "loisirs": "Le programme de vos prochaines sorties.",
    "generique": "Découvrez notre activité en un coup d'œil.",
}

SECTIONS_GROUPE = {
    "restauration": ("Notre menu", "Galerie photo", "Nous trouver"),
    "beaute": ("Nos prestations", "Nos tarifs", "Prendre rendez-vous"),
    "mode": ("Notre catalogue", "Nouveautés", "Commander"),
    "informatique": ("Nos produits", "Demander un devis", "Nous contacter"),
    "alimentation": ("Nos rayons", "Nos promotions", "Nous trouver"),
    "sante": ("Nos services", "Nos horaires", "Nous contacter"),
    "hebergement": ("Nos chambres", "Nos tarifs", "Réserver"),
    "auto": ("Nos prestations", "Nos tarifs", "Prendre rendez-vous"),
    "services": ("Nos services", "À propos", "Nous contacter"),
    "loisirs": ("Notre programme", "Nos horaires", "Nous contacter"),
    "generique": ("Nos services", "À propos", "Nous contacter"),
}

# Chaque famille n'a pas les mêmes besoins : un restaurant vend un menu, un
# institut vend des créneaux, une boutique vend un catalogue. Un seul gabarit
# recoloré ne le montrait pas — la mise en page du cœur de page varie donc
# selon le type d'activité, pas seulement sa couleur.
MISE_EN_PAGE_GROUPE = {
    "restauration": "menu",
    "beaute": "rdv", "sante": "rdv", "auto": "rdv", "hebergement": "rdv", "services": "rdv",
    "mode": "catalogue", "alimentation": "catalogue", "informatique": "catalogue", "generique": "catalogue",
    "loisirs": "agenda",
}

# Un lounge ou un hôtel ne se vend pas avec les mêmes codes visuels qu'une
# clinique ou une supérette : le thème (clair/sombre) suit donc le métier,
# plutôt que d'être imposé à tout le monde.
THEME_GROUPE = {
    "loisirs": "sombre", "hebergement": "sombre", "auto": "sombre",
}

VARS_THEME = {
    "clair": """--accent-clair:color-mix(in srgb, var(--accent) 12%, white);
    --accent-fonce:color-mix(in srgb, var(--accent) 65%, black);
    --bg:#ffffff; --bg-alt:#f7f8fa; --carte-bg:#ffffff; --nav-bg:rgba(255,255,255,.88);
    --texte:#14171f; --doux:#6b7280; --bord:#e8eaed;
    --banniere-bg:#fffbea; --banniere-texte:#8a6416; --banniere-bord:#f3e3ae;""",
    "sombre": """--accent-clair:color-mix(in srgb, var(--accent) 22%, black);
    --accent-fonce:color-mix(in srgb, var(--accent) 40%, white);
    --bg:#0b0d12; --bg-alt:#12151c; --carte-bg:#141821; --nav-bg:rgba(11,13,18,.85);
    --texte:#f1f2f4; --doux:#9aa1ad; --bord:#232a36;
    --banniere-bg:#221d0a; --banniere-texte:#e8c874; --banniere-bord:#3a2f0c;""",
}

ELEMENTS_GROUPE = {
    "restauration": ("Plat du jour", "Boisson fraîche", "Dessert maison"),
    "beaute": ("Coupe & brushing", "Soin du visage", "Manucure"),
    "mode": ("Nouvelle collection", "Meilleure vente", "Édition limitée"),
    "informatique": ("Réparation express", "Accessoires", "Configuration sur mesure"),
    "alimentation": ("Produits frais", "Promotion du jour", "Nouveauté en rayon"),
    "sante": ("Consultation", "Suivi", "Urgence"),
    "hebergement": ("Chambre standard", "Chambre premium", "Suite"),
    "auto": ("Vidange", "Lavage complet", "Diagnostic"),
    "services": ("Premier rendez-vous", "Suivi de dossier", "Accompagnement"),
    "loisirs": ("Soirée à thème", "Cours collectif", "Événement du mois"),
    "generique": ("Service 1", "Service 2", "Service 3"),
}

# ---------------------------------------------------------------------- zones

REPERES = {
    "Kaloum / Almamya": (9.5100, -13.7120),
    "Coleah / Matoto sud": (9.5320, -13.6800),
    "Madina": (9.5470, -13.6700),
    "Matam": (9.5480, -13.6540),
    "Taouyah / Ratoma": (9.5850, -13.6600),
    "Kipé / Kaporo": (9.6050, -13.6450),
    "Nongo / Lambanyi": (9.6250, -13.6250),
    "Cité Enco 5": (9.6300, -13.5930),
    "Sonfonia": (9.6700, -13.5700),
    "Kagbelen / est": (9.6600, -13.5300),
}


def zone_la_plus_proche(lat, lon):
    def dist(p):
        dy = (lat - p[0]) * 111.0
        dx = (lon - p[1]) * 111.0 * math.cos(math.radians(lat))
        return math.hypot(dx, dy)

    return min(REPERES.items(), key=lambda kv: dist(kv[1]))[0]


def distance_km(a_lat, a_lon, b_lat, b_lon):
    dy = (a_lat - b_lat) * 111.0
    dx = (a_lon - b_lon) * 111.0 * math.cos(math.radians(a_lat))
    return math.hypot(dx, dy)


# ------------------------------------------------------------------ nettoyage

def _numero_depuis_chiffres(chiffres):
    """Complète et valide une seule suite de chiffres. Renvoie None si ça ne
    ressemble à aucun numéro guinéen exploitable."""
    if chiffres.startswith("00224"):
        chiffres = chiffres[2:]
    elif len(chiffres) == 9 and chiffres.startswith("6"):
        chiffres = "224" + chiffres
    if len(chiffres) == 12 and chiffres.startswith("224") and chiffres[3] == "6":
        return chiffres
    return None


def normaliser_telephones(brut):
    """
    Renvoie les numéros au format WhatsApp `224XXXXXXXXX`.

    OSM mélange les écritures : `+224622182121`, `00224620883811`,
    `+224 669335443`, `626636629` (sans indicatif), et parfois deux numéros
    dans un même champ (`620585080 / 664243429`, mais aussi parfois séparés
    par un simple espace sans slash : `620585080 664243429`). Sans ce
    nettoyage, l'ouverture du lien WhatsApp échoue silencieusement.
    """
    numeros = []
    for morceau in re.split(r"[/;,]", brut or ""):
        numero = _numero_depuis_chiffres(re.sub(r"\D", "", morceau))
        trouves = [numero] if numero else []

        if not trouves and " " in morceau.strip():
            # Repli seulement si la lecture directe a échoué : un numéro
            # correctement écrit avec des espaces internes (ex. "+224 611 76
            # 85 52") est déjà capturé ci-dessus et n'atteint jamais ce repli
            # — ça évite de le casser en le redécoupant à tort.
            trouves = [n for n in (
                _numero_depuis_chiffres(re.sub(r"\D", "", sous_morceau))
                for sous_morceau in morceau.split()
            ) if n]

        for n in trouves:
            if n not in numeros:
                numeros.append(n)
    return numeros


def image_osm(tags):
    """Une vraie photo, quand OSM en référence une — jamais de photo Google
    Maps (données propriétaires, hors de portée sans clé API payante, ce que
    ce programme s'interdit). Deux sources OSM possibles : un lien direct
    (`image=https://…`), ou une référence Wikimedia Commons (`wikimedia_commons
    =File:xxx.jpg`), convertie via Special:FilePath — le mécanisme standard de
    Commons pour obtenir l'image elle-même à partir de son nom de fichier."""
    image = tags.get("image", "")
    if image.startswith(("http://", "https://")):
        return image
    commons = tags.get("wikimedia_commons", "")
    if commons.startswith("File:"):
        return f"https://commons.wikimedia.org/wiki/Special:FilePath/{urllib.parse.quote(commons[5:])}?width=900"
    return ""


def cle_nom(nom):
    """Nom réduit à l'essentiel, pour repérer « City Market » et « City market »."""
    sans_accents = unicodedata.normalize("NFD", nom)
    sans_accents = "".join(c for c in sans_accents if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", sans_accents.lower())


# ------------------------------------------------------------------- Overpass

def interroger_overpass(ville="Conakry"):
    clauses = "\n".join(
        f'  nwr[{f}][!"website"][!"contact:website"]'
        f'[~"^(phone|contact:phone)$"~"."](area.zone);'
        for f in FILTRES_OSM
    )
    requete = f'[out:json][timeout:120];\narea["name"="{ville}"]->.zone;\n(\n{clauses}\n);\nout center tags;'

    # Depuis avril 2026, overpass-api.de exige un User-Agent explicite et
    # renvoie 406 s'il est absent. La politique d'usage d'Overpass demande
    # d'identifier son outil plutôt que d'imiter un navigateur — on s'y tient.
    entetes = {"User-Agent": "ProspectionVisionAmah/1.0 (usage manuel, sans automatisation d'envoi)"}

    ESSAIS_PAR_SERVEUR = 2
    ATTENTE_SECONDES = 15  # 504 = serveur surchargé, pas une erreur de requête : on laisse retomber la charge.

    derniere_erreur = None
    for serveur in OVERPASS_ENDPOINTS:
        for essai in range(1, ESSAIS_PAR_SERVEUR + 1):
            print(f"Interrogation d'OpenStreetMap pour {ville} ({serveur})… (essai {essai}/{ESSAIS_PAR_SERVEUR})")
            try:
                reponse = requests.post(serveur, data={"data": requete}, headers=entetes, timeout=180)
            except requests.exceptions.RequestException as e:
                derniere_erreur = str(e)
                print(f"[!] Connexion impossible : {e}")
                time.sleep(ATTENTE_SECONDES)
                continue

            if reponse.status_code == 200:
                return reponse.json().get("elements", [])

            derniere_erreur = f"{reponse.status_code} :\n{reponse.text[:400]}"
            if reponse.status_code == 504:
                # Serveur surchargé, pas notre requête qui est en cause (elle passe
                # sans problème sur overpass-turbo.eu) : on retente, puis on bascule.
                print(f"[!] Overpass a répondu 504 (serveur surchargé), nouvelle tentative dans {ATTENTE_SECONDES}s…")
                time.sleep(ATTENTE_SECONDES)
                continue
            else:
                # Erreur non liée à la charge (ex. requête invalide) : inutile de
                # réessayer sur ce serveur, ni sur le suivant.
                print(f"[!] Overpass a répondu {derniere_erreur}")
                sys.exit(1)

    print(f"[!] Impossible d'obtenir une réponse d'OpenStreetMap après plusieurs tentatives : {derniere_erreur}")
    print("Les serveurs Overpass sont probablement surchargés en ce moment. Réessaie dans quelques minutes.")
    sys.exit(1)


# -------------------------------------------------------------------- filtres

def extraire(elements):
    fiches = []
    rejets = {"sans_nom": 0, "sans_numero_valide": 0}

    for e in elements:
        tags = e.get("tags", {})
        nom = (tags.get("name") or "").strip()
        if not nom:
            rejets["sans_nom"] += 1
            continue

        numeros = normaliser_telephones(tags.get("phone") or tags.get("contact:phone"))
        if not numeros:
            rejets["sans_numero_valide"] += 1
            continue

        centre = e.get("center") or {}
        lat = e.get("lat", centre.get("lat"))
        lon = e.get("lon", centre.get("lon"))
        if lat is None or lon is None:
            continue

        type_osm = (tags.get("shop") or tags.get("amenity") or tags.get("tourism")
                    or tags.get("office") or tags.get("craft") or tags.get("leisure") or "")
        fiches.append({
            "id_osm": f'{e["type"]}/{e["id"]}',
            "nom": nom,
            "numeros": numeros,
            "lat": lat,
            "lon": lon,
            "type_osm": type_osm,
            "famille": famille(type_osm),
            "besoin": BESOIN.get(famille(type_osm), "site"),
            "email": tags.get("email", ""),
            "reseau": tags.get("facebook") or tags.get("contact:facebook")
                      or tags.get("contact:instagram", ""),
            "rue": tags.get("addr:street", ""),
            "image": image_osm(tags),
            # website:menu / contact:menu : lien vers un menu existant. Le tag
            # "menu" seul n'est PAS ça en OSM — c'est une description du type
            # de cuisine (ex. "regional"), pas une URL ; on ne le reprend pas.
            "menu_url": tags.get("website:menu") or tags.get("contact:menu") or "",
        })
    return fiches, rejets


def dedupliquer(fiches):
    """
    Deux passes, car OSM produit deux types de doublons distincts :

    1. Le même établissement saisi plusieurs fois (parfois des dizaines de fois,
       à quelques mètres d'écart) — repéré par nom identique et proximité.
    2. Plusieurs points de vente partageant un unique numéro (une chaîne).
       Les contacter tous, c'est écrire plusieurs fois à la même personne.
    """
    retenues, doublons = [], []

    for f in fiches:
        jumeau = next(
            (g for g in retenues
             if cle_nom(g["nom"]) == cle_nom(f["nom"])
             and distance_km(g["lat"], g["lon"], f["lat"], f["lon"]) < 0.15),
            None,
        )
        if jumeau:
            doublons.append((f["nom"], "même nom, même endroit"))
            continue
        retenues.append(f)

    finales, vus_tel = [], {}
    for f in retenues:
        # Comparé sur TOUS les numéros de la fiche, pas seulement le premier :
        # deux fiches peuvent partager un numéro qui n'est pas en première
        # position sur l'une d'elles (ex. A=[X,Y], B=[Y,Z]).
        partage = next((n for n in f["numeros"] if n in vus_tel), None)
        if partage:
            doublons.append((f["nom"], f'même numéro que « {vus_tel[partage]} »'))
            continue
        for n in f["numeros"]:
            vus_tel[n] = f["nom"]
        finales.append(f)

    return finales, doublons


def score(f):
    """Un commerce déjà présent en ligne comprend l'offre plus vite."""
    return (1 if f["email"] else 0) + (1 if f["reseau"] else 0)


# --------------------------------------------------------------- persistance

def charger_fiches_connues():
    """Ce fichier stocke les fiches complètes (pas seulement leurs identifiants) :
    prospects.csv/html/pdf doivent réafficher TOUT l'historique à chaque
    lancement, pas seulement les nouveautés du jour — sinon le suivi de
    statut, les relances et les stats de prospects.html perdraient tout ce qui
    a été trouvé lors d'exécutions précédentes dès le lancement suivant
    (la tâche planifiée hebdomadaire, en particulier)."""
    if not os.path.exists(FICHIER_DEJA_VUS):
        return {}
    with open(FICHIER_DEJA_VUS, encoding="utf-8") as f:
        donnees = json.load(f)
    if isinstance(donnees, list):
        # Ancien format (une simple liste d'identifiants, sans le détail des
        # fiches) : impossible à reconstruire, on repart proprement plutôt
        # que d'afficher des fiches à moitié vides.
        return {}
    return donnees


def sauvegarder_fiches_connues(fiches_par_id):
    f, _ = ouvrir_en_ecriture(FICHIER_DEJA_VUS, encoding="utf-8")
    with f:
        json.dump(fiches_par_id, f, ensure_ascii=False, indent=1)


def chemin_de_repli(chemin):
    """Si le fichier visé est verrouillé (ouvert dans Excel, par ex.), on écrit
    à côté sous un nom horodaté plutôt que de planter le programme."""
    racine, ext = os.path.splitext(chemin)
    return f"{racine}-{time.strftime('%Y%m%d-%H%M%S')}{ext}"


def ouvrir_en_ecriture(chemin, **kwargs):
    """open() avec repli si le fichier — ou le dossier lui-même — est bloqué en écriture."""
    try:
        return open(chemin, "w", **kwargs), chemin
    except PermissionError:
        pass

    # Un fichier verrouillé (ouvert dans Excel) empêche seulement CE nom précis :
    # on retente avec un nom horodaté, dans le même dossier.
    repli = chemin_de_repli(chemin)
    try:
        f = open(repli, "w", **kwargs)
        print(
            f"[!] Impossible d'écrire dans « {chemin} » (probablement ouvert dans "
            f"Excel ou un autre programme). Écriture dans « {repli} » à la place — "
            f"ferme le fichier original avant la prochaine exécution."
        )
        return f, repli
    except PermissionError:
        pass

    # Si même un nom tout neuf est refusé dans ce dossier, ce n'est plus un
    # fichier verrouillé mais le dossier lui-même qui est protégé — le cas
    # classique est le « Contrôle d'accès aux dossiers » de Windows Defender
    # (Sécurité Windows → Protection contre les rançongiciels), qui bloque par
    # défaut l'écriture dans Bureau/Documents/Images pour les programmes non
    # autorisés. On écrit alors dans le dossier applicatif (toujours accessible),
    # sous le même nom à chaque fois — pas d'horodatage, pour ne pas s'y perdre.
    repli_app = os.path.join(DOSSIER_APP, os.path.basename(chemin))
    print(
        f"[!] Le dossier « {os.path.dirname(os.path.abspath(chemin)) or '.'} » refuse "
        f"aussi l'écriture d'un nouveau fichier — ce n'est probablement pas Excel, mais "
        f"le Contrôle d'accès aux dossiers de Windows (Sécurité Windows → Protection "
        f"contre les rançongiciels) qui bloque Python dans ce dossier. Écriture dans "
        f"« {repli_app} » à la place. Pour écrire directement sur le Bureau la "
        f"prochaine fois : autorise python.exe dans cette protection, ou désactive-la."
    )
    return open(repli_app, "w", **kwargs), repli_app


# ------------------------------------------------------------ journal d'envois

COLONNES_JOURNAL = ["date", "nom", "zone", "categorie", "telephone", "statut", "message"]


def fusionner_journal_envois():
    """Absorbe l'export produit par le bouton « Exporter le journal » de
    prospects.html dans un classeur Excel dédié, géré entièrement par ce
    script — jamais dans le fichier de suivi manuel de l'utilisateur (qui
    contient déjà ses propres notes et mises en forme, qu'un aller-retour
    automatisé via openpyxl risquerait d'abîmer).

    Renvoie None si aucun export n'est présent, sinon le nombre de nouvelles
    lignes ajoutées (0 si l'export ne contenait rien de plus que ce qui est
    déjà dans le classeur — le bouton réexporte tout à chaque fois, pas
    seulement les nouveautés)."""
    if not os.path.exists(FICHIER_JOURNAL_EXPORT):
        return None

    import openpyxl

    with open(FICHIER_JOURNAL_EXPORT, encoding="utf-8-sig", newline="") as f:
        lignes_export = list(csv.DictReader(f, delimiter=";"))
    if not lignes_export:
        return 0

    if os.path.exists(FICHIER_JOURNAL_XLSX):
        classeur = openpyxl.load_workbook(FICHIER_JOURNAL_XLSX)
        feuille = classeur.active
        # (telephone, date) identifie une entrée de façon unique : comme
        # l'export contient tout l'historique à chaque fois (pas seulement
        # les nouveautés), c'est ce qui évite de dupliquer les lignes déjà
        # importées lors d'une fusion précédente.
        deja = {(ligne[4], ligne[0]) for ligne in feuille.iter_rows(min_row=2, values_only=True) if ligne[0]}
    else:
        classeur = openpyxl.Workbook()
        feuille = classeur.active
        feuille.title = "Journal des envois"
        feuille.append(["Date", "Nom", "Zone", "Categorie", "Telephone", "Statut", "Message envoyé"])
        deja = set()

    ajoutees = 0
    for ligne in lignes_export:
        cle = (ligne.get("telephone", ""), ligne.get("date", ""))
        if cle in deja:
            continue
        feuille.append([ligne.get(c, "") for c in COLONNES_JOURNAL])
        deja.add(cle)
        ajoutees += 1

    if not ajoutees:
        return 0

    for cible in (FICHIER_JOURNAL_XLSX, chemin_de_repli(FICHIER_JOURNAL_XLSX),
                  os.path.join(DOSSIER_APP, os.path.basename(FICHIER_JOURNAL_XLSX))):
        try:
            classeur.save(cible)
            if cible != FICHIER_JOURNAL_XLSX:
                print(f"[!] Impossible d'écrire dans « {FICHIER_JOURNAL_XLSX} » — écrit dans « {cible} » à la place.")
            return ajoutees
        except PermissionError:
            continue
    raise PermissionError(
        f"Impossible d'écrire {FICHIER_JOURNAL_XLSX} : dossier bloqué en écriture, y compris {DOSSIER_APP}."
    )


def ecrire_csv(fiches, chemin=FICHIER_CSV):
    """Export de secours, pour classement ou sauvegarde hors-ligne."""
    colonnes = ["zone", "nom", "categorie", "besoin", "telephone", "lien_whatsapp",
                "emplacement", "rue", "email", "reseau", "priorite",
                "statut", "message", "demo"]
    f, chemin_reel = ouvrir_en_ecriture(chemin, newline="", encoding="utf-8-sig")
    with f:
        w = csv.writer(f, delimiter=";")
        w.writerow(colonnes)
        for p in fiches:
            principal = p["numeros"][0]
            w.writerow([
                p["zone"], p["nom"], p["type_osm"],
                p.get("besoin", "site"),
                " / ".join("+" + n for n in p["numeros"]),
                f"https://wa.me/{principal}",
                f'https://www.google.com/maps?q={p["lat"]},{p["lon"]}',
                p["rue"], p["email"], p["reseau"],
                "haute" if score(p) else "normale",
                "à contacter", p["message"], p.get("demo", ""),
            ])
    # `utf-8-sig` et `;` : Excel en configuration française ouvre correctement.
    return chemin_reel


# ------------------------------------------------------------ interface HTML

GABARIT_HTML = """<!doctype html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Prospection — Vision Amah</title>
<style>
  :root { --bg:#050505; --carte:#0e1118; --bord:#1e2532; --accent:#3b82f6; --texte:#e8eaed; --doux:#8b93a1; }
  * { box-sizing:border-box; }
  body { margin:0; background:var(--bg); color:var(--texte);
         font-family:system-ui,-apple-system,"Segoe UI",Roboto,sans-serif; }
  header { position:sticky; top:0; z-index:10; background:rgba(5,5,5,.96);
           border-bottom:1px solid var(--bord); padding:14px 16px; backdrop-filter:blur(8px); }
  h1 { margin:0 0 10px; font-size:15px; text-transform:uppercase; letter-spacing:.18em; }
  .compteur { color:var(--doux); font-size:13px; font-weight:normal; letter-spacing:0; text-transform:none; }
  .barre { display:flex; gap:8px; flex-wrap:wrap; }
  input,select { background:#12161f; border:1px solid var(--bord); color:var(--texte);
                 padding:9px 12px; border-radius:9px; font-size:14px; flex:1; min-width:130px; }
  main { padding:16px; max-width:820px; margin:0 auto; }
  .stats-panel { display:none; margin:0 16px 14px; border:1px solid var(--bord); border-radius:12px;
                 padding:12px 14px; background:#0b0e15; overflow-x:auto; }
  .stats-panel.ouvert { display:block; }
  table.stats { width:100%; border-collapse:collapse; font-size:12.5px; white-space:nowrap; }
  table.stats th, table.stats td { text-align:left; padding:6px 10px; border-bottom:1px solid var(--bord); }
  table.stats th { color:var(--doux); font-weight:600; }
  .section { margin-bottom:14px; border:1px solid var(--bord); border-radius:14px; overflow:hidden; }
  .entete {
    display:flex; justify-content:space-between; align-items:center; gap:10px;
    padding:14px 16px; background:#0b0e15; cursor:pointer; user-select:none;
  }
  .entete .titre { font-weight:650; font-size:14.5px; }
  .entete .sous { color:var(--doux); font-size:12.5px; font-weight:normal; }
  .fleche { color:var(--doux); font-size:12px; transition:transform .15s; }
  .section.repliee .fleche { transform:rotate(-90deg); }
  .corps { display:grid; gap:10px; padding:12px; }
  .section.repliee .corps { display:none; }
  .fiche { background:var(--carte); border:1px solid var(--bord); border-radius:12px; padding:14px; }
  .fiche.statut-contacte, .fiche.statut-pas_interesse { opacity:.55; }
  .fiche.statut-signe { opacity:.7; border-color:#3a2f0c; }
  .fiche.statut-repondu { border-color:#1f4a2c; }
  .fiche.relance { opacity:1 !important; border-color:#7f1d1d; }
  .haut { display:flex; justify-content:space-between; gap:10px; align-items:flex-start; }
  .nom { font-weight:650; font-size:15.5px; }
  .meta { color:var(--doux); font-size:12.5px; margin-top:3px; }
  .etiquettes { display:flex; gap:6px; flex-direction:column; align-items:flex-end; }
  .etiq { font-size:10px; text-transform:uppercase; letter-spacing:.12em;
          padding:3px 9px; border-radius:99px; white-space:nowrap; }
  .prio { background:rgba(59,130,246,.16); color:#7fb0ff; }
  .relance { background:rgba(244,63,94,.18); color:#fb7185; }
  .etat-contacte { background:rgba(59,130,246,.16); color:#7fb0ff; }
  .etat-repondu { background:rgba(34,197,94,.18); color:#4ade80; }
  .etat-signe { background:rgba(234,179,8,.18); color:#fbbf24; }
  .etat-pas_interesse { background:rgba(139,147,161,.18); color:#8b93a1; }
  .actions { display:flex; gap:8px; flex-wrap:wrap; margin-top:12px; }
  a.btn, button.btn, select.btn { flex:1; min-width:112px; text-align:center; text-decoration:none;
        padding:11px 10px; border-radius:9px; font-size:13.5px; font-weight:600;
        border:1px solid var(--bord); background:#151a24; color:var(--texte); cursor:pointer; }
  a.wa { background:var(--accent); border-color:var(--accent); color:#fff; }
  .msg { display:none; margin-top:11px; white-space:pre-wrap; font-size:13px;
         line-height:1.55; color:#c7ccd6; background:#0a0d13;
         border:1px solid var(--bord); border-radius:9px; padding:12px; }
  .msg.ouvert { display:block; }
  .vide { text-align:center; color:var(--doux); padding:40px 0; }
  .badge-auto { background:#78350f22; color:#f59e0b; border:1px solid #f59e0b55; }
  .badge-site { background:#1e3a5f22; color:#60a5fa; border:1px solid #3b82f655; }
</style>
</head>
<body>
<header>
  <h1>Prospection Vision&nbsp;Amah <span class="compteur" id="compteur"></span></h1>
  <div class="barre">
    <input id="q" placeholder="Rechercher un nom…">
    <select id="zone"><option value="">Toutes les zones</option>__ZONES__</select>
    <select id="besoin">
      <option value="">⚡ Site + Auto</option>
      <option value="automatisation">⚡ Automatisation</option>
      <option value="site">🌐 Site web</option>
    </select>
    <select id="etat">
      <option value="restants">À contacter</option>
      <option value="tous">Tous</option>
      <option value="prio">Prioritaires</option>
      <option value="relance">À relancer</option>
      <option value="contactes">Contactés (tous)</option>
      <option value="repondus">Ont répondu</option>
      <option value="signes">Signés</option>
      <option value="pas_interesses">Pas intéressés</option>
    </select>
    <button class="btn" id="bouton-stats" type="button">📊 Stats</button>
    <button class="btn" id="bouton-export" type="button">💾 Exporter le journal</button>
  </div>
  <div class="stats-panel" id="stats-panel"></div>
</header>
<main id="liste"></main>

<script>
const DONNEES = __DONNEES__;
const LABELS = __LABELS__;
const ORDRE = __ORDRE__;
const ORDRE_AUTO = __ORDRE_AUTO__;
const ORDRE_SITE = __ORDRE_SITE__;

const JOUR_MS = 24 * 60 * 60 * 1000;
const DELAI_RELANCE_JOURS = 5;

// L'ordre des clés fixe l'ordre des options du menu déroulant par fiche.
const STATUT_LABELS = {
  a_contacter: "À contacter",
  contacte: "Contacté",
  repondu: "A répondu",
  signe: "Signé",
  pas_interesse: "Pas intéressé",
};

const CLE_STATUTS = "va_statuts";
const CLE_REPLIEES = "va_categories_repliees";
// "a_contacter" n'est jamais stocké : son absence de statuts EST l'état par
// défaut, ce qui évite d'avoir à migrer l'ancien CLE_FAITS (booléen) séparément.
let statuts = JSON.parse(localStorage.getItem(CLE_STATUTS) || "{}");
let repliees = new Set(JSON.parse(localStorage.getItem(CLE_REPLIEES) || "[]"));

function etatDe(id) {
  return (statuts[id] && statuts[id].etat) || "a_contacter";
}

function estARelancer(id) {
  const s = statuts[id];
  if (!s || s.etat !== "contacte") return false;
  return (Date.now() - s.date) >= DELAI_RELANCE_JOURS * JOUR_MS;
}

function definirStatut(id, etat) {
  if (etat === "a_contacter") {
    delete statuts[id];
  } else {
    statuts[id] = { etat, date: Date.now() };
  }
  localStorage.setItem(CLE_STATUTS, JSON.stringify(statuts));
}

function basculerSection(groupe) {
  repliees.has(groupe) ? repliees.delete(groupe) : repliees.add(groupe);
  localStorage.setItem(CLE_REPLIEES, JSON.stringify([...repliees]));
  document.querySelector(`.section[data-groupe="${groupe}"]`)
    .classList.toggle("repliee", repliees.has(groupe));
}

async function copier(id, bouton) {
  const p = DONNEES.find(d => d.id === id);
  try {
    await navigator.clipboard.writeText(p.message);
    bouton.textContent = "✓ Copié";
  } catch {
    bouton.closest(".fiche").querySelector(".msg").classList.add("ouvert");
    bouton.textContent = "Sélectionne et copie ↓";
  }
  setTimeout(() => (bouton.textContent = "Copier le message"), 2200);
}

function majCompteur() {
  const t = DONNEES.length;
  const restants = DONNEES.filter(d => etatDe(d.id) === "a_contacter").length;
  document.getElementById("compteur").textContent = `— ${restants} restants sur ${t}`;
}

function construireFiche(p) {
  const etat = etatDe(p.id);
  const relance = estARelancer(p.id);
  const el = document.createElement("div");
  el.className = `fiche statut-${etat}` + (relance ? " relance" : "");
  el.innerHTML = `
    <div class="haut">
      <div>
        <div class="nom"></div>
        <div class="meta"></div>
      </div>
      <div class="etiquettes">
        ${p.besoin === "automatisation" ? '<span class="etiq badge-auto">⚡ Auto</span>' : '<span class="etiq badge-site">🌐 Site</span>'}
        ${p.prio ? '<span class="etiq prio">prioritaire</span>' : ""}
        ${relance ? '<span class="etiq relance">à relancer</span>' : ""}
        ${etat !== "a_contacter" ? `<span class="etiq etat-${etat}"></span>` : ""}
      </div>
    </div>
    <div class="actions">
      <a class="btn wa" href="${p.wa}" target="_blank" rel="noopener">WhatsApp</a>
      <a class="btn" href="${p.carte}" target="_blank" rel="noopener">Voir sur la carte</a>
      ${p.demo ? `<a class="btn" href="${p.demo}" target="_blank" rel="noopener">Voir la démo</a>` : ""}
      <button class="btn copier">Copier le message</button>
    </div>
    <div class="actions">
      <select class="btn statut-select"></select>
    </div>
    <div class="msg"></div>`;
  // textContent : les noms viennent d'OSM, donc de contributeurs tiers —
  // on ne les injecte jamais en HTML.
  el.querySelector(".nom").textContent = p.nom;
  el.querySelector(".meta").textContent = `${p.zone} · ${p.cat} · ${p.tel}`;
  el.querySelector(".msg").textContent = p.message;
  const badgeEtat = el.querySelector(`.etat-${etat}`);
  if (badgeEtat) badgeEtat.textContent = STATUT_LABELS[etat];
  const select = el.querySelector(".statut-select");
  for (const [valeur, libelle] of Object.entries(STATUT_LABELS)) {
    const option = document.createElement("option");
    option.value = valeur;
    option.textContent = libelle;
    option.selected = valeur === etat;
    select.appendChild(option);
  }
  el.querySelector(".copier").onclick = ev => copier(p.id, ev.currentTarget);
  select.onchange = ev => {
    definirStatut(p.id, ev.currentTarget.value);
    majCompteur();
    afficher();
  };
  return el;
}

function afficher() {
  const q = document.getElementById("q").value.trim().toLowerCase();
  const z = document.getElementById("zone").value;
  const e = document.getElementById("etat").value;
  const b = document.getElementById("besoin").value;

  const vus = DONNEES.filter(p => {
    if (q && !p.nom.toLowerCase().includes(q)) return false;
    if (z && p.zone !== z) return false;
    if (b && p.besoin !== b) return false;
    const etat = etatDe(p.id);
    if (e === "restants" && etat !== "a_contacter") return false;
    if (e === "prio" && !p.prio) return false;
    if (e === "relance" && !estARelancer(p.id)) return false;
    if (e === "contactes" && etat === "a_contacter") return false;
    if (e === "repondus" && etat !== "repondu") return false;
    if (e === "signes" && etat !== "signe") return false;
    if (e === "pas_interesses" && etat !== "pas_interesse") return false;
    return true;
  });

  const liste = document.getElementById("liste");
  liste.innerHTML = "";

  if (!vus.length) {
    liste.innerHTML = '<p class="vide">Aucun prospect ne correspond.</p>';
    return;
  }

  function rendreBloc(ordre, titre, couleur) {
    let affiche = false;
    for (const groupe of ordre) {
      const items = vus.filter(p => p.groupe === groupe);
      if (!items.length) continue;
      if (!affiche) {
        const entete = document.createElement("div");
        entete.className = "entete-bloc";
        entete.style.cssText = `border-left:3px solid ${couleur};padding:6px 10px;margin:18px 0 6px;font-size:12px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;color:${couleur};`;
        entete.textContent = titre;
        liste.appendChild(entete);
        affiche = true;
      }
      const section = document.createElement("div");
      section.className = "section" + (repliees.has(groupe) ? " repliee" : "");
      section.dataset.groupe = groupe;
      section.innerHTML = `
        <div class="entete">
          <span class="titre">${LABELS[groupe]} <span class="sous">— ${items.length}</span></span>
          <span class="fleche">▾</span>
        </div>
        <div class="corps"></div>`;
      section.querySelector(".entete").onclick = () => basculerSection(groupe);
      const corps = section.querySelector(".corps");
      for (const p of items) corps.appendChild(construireFiche(p));
      liste.appendChild(section);
    }
  }

  rendreBloc(ORDRE_AUTO, "⚡ Automatisation — processus à automatiser", "#f59e0b");
  rendreBloc(ORDRE_SITE, "🌐 Site web — présence en ligne", "#3b82f6");
}

function construireStats() {
  const lignes = ORDRE.map(groupe => {
    const items = DONNEES.filter(p => p.groupe === groupe);
    if (!items.length) return null;
    const contactes = items.filter(p => etatDe(p.id) !== "a_contacter").length;
    const repondus = items.filter(p => ["repondu", "signe"].includes(etatDe(p.id))).length;
    const signes = items.filter(p => etatDe(p.id) === "signe").length;
    const taux = contactes ? Math.round((repondus / contactes) * 100) : 0;
    return `<tr><td>${LABELS[groupe]}</td><td>${items.length}</td><td>${contactes}</td>`
         + `<td>${repondus}</td><td>${signes}</td><td>${taux}%</td></tr>`;
  }).filter(Boolean).join("");

  document.getElementById("stats-panel").innerHTML = `
    <table class="stats">
      <thead><tr><th>Catégorie</th><th>Total</th><th>Contactés</th><th>Répondu</th><th>Signé</th><th>Taux de réponse</th></tr></thead>
      <tbody>${lignes}</tbody>
    </table>`;
}

document.getElementById("bouton-stats").onclick = () => {
  const panneau = document.getElementById("stats-panel");
  if (panneau.classList.toggle("ouvert")) construireStats();
};

// Une page statique ne peut pas écrire de fichier toute seule (sécurité du
// navigateur) : ce bouton exporte donc un CSV — que prospect_finder.py
// absorbe automatiquement dans un classeur Excel dédié au lancement suivant.
// Il reprend TOUT l'historique à chaque fois (pas que les nouveautés) : la
// fusion côté script se charge de ne pas dupliquer ce qui est déjà importé.
async function exporterJournal() {
  const lignes = DONNEES
    .filter(p => etatDe(p.id) !== "a_contacter")
    .map(p => {
      const s = statuts[p.id];
      return {
        date: new Date(s.date).toISOString(),
        nom: p.nom, zone: p.zone, categorie: p.cat, telephone: p.tel,
        statut: STATUT_LABELS[etatDe(p.id)], message: p.message,
      };
    });

  if (!lignes.length) {
    alert("Aucun prospect avec un statut renseigné pour le moment — rien à exporter.");
    return;
  }

  const colonnes = ["date", "nom", "zone", "categorie", "telephone", "statut", "message"];
  const echapper = v => `"${String(v).replace(/"/g, '""')}"`;
  const contenu = "﻿" + colonnes.join(";") + "\\n"
    + lignes.map(l => colonnes.map(c => echapper(l[c])).join(";")).join("\\n");
  const nomFichier = "journal_envois.csv";

  if (window.showSaveFilePicker) {
    try {
      const poignee = await window.showSaveFilePicker({
        suggestedName: nomFichier,
        types: [{ description: "CSV", accept: { "text/csv": [".csv"] } }],
      });
      const flux = await poignee.createWritable();
      await flux.write(contenu);
      await flux.close();
      return;
    } catch (e) {
      if (e.name === "AbortError") return;
      // sinon on retombe sur le téléchargement classique ci-dessous
    }
  }

  const lien = document.createElement("a");
  lien.href = URL.createObjectURL(new Blob([contenu], { type: "text/csv" }));
  lien.download = nomFichier;
  lien.click();
  URL.revokeObjectURL(lien.href);
}

document.getElementById("bouton-export").onclick = exporterJournal;

for (const id of ["q", "zone", "besoin", "etat"]) {
  document.getElementById(id).addEventListener("input", afficher);
}
majCompteur();
afficher();
</script>
</body>
</html>
"""


def ecrire_html(fiches, chemin=FICHIER_HTML):
    donnees = []
    for p in fiches:
        principal = p["numeros"][0]
        donnees.append({
            "id": f'{p["nom"]}|{principal}',
            "nom": p["nom"],
            "zone": p["zone"],
            "cat": p["type_osm"],
            "groupe": p["famille"],
            "tel": " / ".join("+" + n for n in p["numeros"]),
            "wa": f"https://wa.me/{principal}",
            "carte": f'https://www.google.com/maps?q={p["lat"]},{p["lon"]}',
            "prio": bool(score(p)),
            "besoin": p.get("besoin", "site"),
            "message": p["message"],
            "demo": f"demos/{pathlib.Path(p['demo']).name}" if p.get("demo") else "",
        })

    zones = sorted({d["zone"] for d in donnees})
    options = "".join(f"<option>{html.escape(z)}</option>" for z in zones)

    # .replace("</", "<\\/") : les noms viennent d'OSM, une base modifiable par
    # n'importe qui. Sans ça, un nom contenant littéralement "</script>"
    # fermerait la balise <script> en avance et casserait la page (ou pire).
    page = (GABARIT_HTML
            .replace("__ZONES__", options)
            .replace("__DONNEES__", json.dumps(donnees, ensure_ascii=False).replace("</", "<\\/"))
            .replace("__LABELS__", json.dumps(LABELS_GROUPE, ensure_ascii=False))
            .replace("__ORDRE__", json.dumps(ORDRE_GROUPES, ensure_ascii=False))
            .replace("__ORDRE_AUTO__", json.dumps(ORDRE_AUTO, ensure_ascii=False))
            .replace("__ORDRE_SITE__", json.dumps(ORDRE_SITE, ensure_ascii=False)))

    f, chemin_reel = ouvrir_en_ecriture(chemin, encoding="utf-8")
    with f:
        f.write(page)
    return chemin_reel


# ------------------------------------------------------------- pages de démo

GABARIT_DEMO = """<!doctype html>
<html lang="fr">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__NOM__</title>
<style>
  :root {
    --accent:__COULEUR__;
    __VARS_THEME__
  }
  * { box-sizing:border-box; }
  body { margin:0; background:var(--bg); color:var(--texte);
         font-family:system-ui,-apple-system,"Segoe UI",Roboto,sans-serif; }
  .banniere { text-align:center; padding:9px 16px; background:var(--banniere-bg);
              color:var(--banniere-texte); font-size:12.5px;
              border-bottom:1px solid var(--banniere-bord); }
  nav { position:sticky; top:0; z-index:10; display:flex; justify-content:space-between;
        align-items:center; padding:16px 24px; background:var(--nav-bg);
        backdrop-filter:blur(8px); border-bottom:1px solid var(--bord); flex-wrap:wrap; gap:10px; }
  nav .marque { display:flex; align-items:center; gap:10px; font-weight:700; font-size:16px; }
  nav .marque .puce { width:30px; height:30px; border-radius:9px; display:flex;
        align-items:center; justify-content:center; font-size:15px;
        background:linear-gradient(135deg, var(--accent), var(--accent-fonce)); }
  nav .liens { display:flex; gap:22px; color:var(--doux); font-size:13.5px; }
  a.btn, button.btn { text-decoration:none; padding:12px 20px; border-radius:10px; font-size:14px;
        font-weight:600; border:1px solid var(--bord); color:var(--texte); background:var(--carte-bg);
        cursor:pointer; display:inline-block; }
  a.principal { background:var(--accent); border-color:var(--accent); color:#fff;
                box-shadow:0 10px 22px -10px color-mix(in srgb, var(--accent) 70%, transparent); }
  .hero { padding:60px 24px 40px; text-align:center;
          background:linear-gradient(180deg, var(--accent-clair), var(--bg) 75%); }
  .photo-hero { width:128px; height:128px; margin:0 auto 22px; border-radius:30px;
        display:flex; align-items:center; justify-content:center; font-size:56px;
        background:linear-gradient(135deg, var(--accent), var(--accent-fonce));
        box-shadow:0 20px 44px -16px color-mix(in srgb, var(--accent) 65%, transparent); }
  .photo-hero.reelle { background-size:cover; background-position:center; font-size:0; }
  .photo-reelle-etiq { display:inline-block; margin-top:-6px; margin-bottom:20px; font-size:11px;
        color:var(--doux); }
  .etiquette { display:inline-block; background:var(--accent-clair); color:var(--accent-fonce);
               font-size:12px; text-transform:uppercase; letter-spacing:.1em;
               padding:5px 14px; border-radius:99px; margin-bottom:14px; font-weight:600; }
  .hero h1 { font-size:clamp(28px,5vw,42px); margin:0 0 12px; letter-spacing:-.01em; }
  .hero p { color:var(--doux); font-size:16px; max-width:480px; margin:0 auto 28px; }
  .cta { display:flex; gap:12px; justify-content:center; flex-wrap:wrap; }
  .confiance { display:flex; gap:18px; justify-content:center; flex-wrap:wrap; margin-top:30px;
               color:var(--doux); font-size:13px; }
  .confiance b { color:var(--texte); }
  main { max-width:820px; margin:0 auto; padding:8px 24px 10px; }
  section { padding:46px 0; }
  section h2 { font-size:22px; margin:0 0 8px; text-align:center; }
  section .sous-titre { text-align:center; color:var(--doux); font-size:14px; margin:0 0 34px; }
  .grille { display:grid; grid-template-columns:repeat(auto-fit,minmax(210px,1fr)); gap:18px; }
  .carte { background:var(--carte-bg); border:1px solid var(--bord); border-radius:16px; padding:24px;
           box-shadow:0 2px 10px -6px rgba(20,23,31,.06); }
  .carte .icone { width:44px; height:44px; border-radius:12px; display:flex; align-items:center;
        justify-content:center; font-size:20px; background:var(--accent-clair); margin-bottom:14px; }
  .carte h3 { margin:0 0 8px; font-size:15.5px; }
  .carte p { margin:0; color:var(--doux); font-size:13.5px; line-height:1.55; }
  .carte.menu { max-width:520px; margin:0 auto; padding:8px 28px; }
  .ligne-menu { display:flex; justify-content:space-between; align-items:baseline;
        padding:16px 0; border-bottom:1px dashed var(--bord); font-size:15.5px; }
  .ligne-menu:last-child { border-bottom:none; }
  .ligne-menu .prix { color:var(--accent-fonce); font-weight:700; font-size:13.5px; white-space:nowrap; }
  .grille.produits { display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr)); gap:18px; }
  .produit { text-align:center; }
  .produit .photo-vide { margin-bottom:12px; }
  .produit h4 { margin:0 0 4px; font-size:14px; }
  .produit .prix { color:var(--accent-fonce); font-weight:700; font-size:13px; }
  .liste-services, .liste-agenda { display:flex; flex-direction:column; gap:12px; max-width:580px; margin:0 auto; }
  .service, .evenement { display:flex; justify-content:space-between; align-items:center; gap:14px;
        background:var(--carte-bg); border:1px solid var(--bord); border-radius:14px; padding:16px 20px;
        box-shadow:0 2px 10px -6px rgba(20,23,31,.06); }
  .service h4, .evenement h4 { margin:0 0 4px; font-size:15px; }
  .service .doux, .evenement .doux { color:var(--doux); font-size:12.5px; }
  a.btn.mini { padding:9px 16px; font-size:12.5px; flex-shrink:0; }
  .evenement { justify-content:flex-start; }
  .evenement .date { background:var(--accent-clair); color:var(--accent-fonce); font-weight:700;
        font-size:12px; text-transform:uppercase; border-radius:10px; padding:10px 14px;
        text-align:center; min-width:56px; flex-shrink:0; }
  .galerie { display:grid; grid-template-columns:repeat(3,1fr); gap:14px; }
  .photo-vide { aspect-ratio:1; border-radius:14px; display:flex; flex-direction:column; gap:6px;
        align-items:center; justify-content:center; background:var(--bg-alt);
        border:1.5px dashed var(--bord); color:var(--doux); font-size:12px; text-align:center; }
  .photo-vide span:first-child { font-size:24px; }
  .avis-client { background:var(--bg-alt); border-radius:18px; padding:32px; text-align:center; }
  .avis-client .etoiles { color:#f59e0b; font-size:18px; letter-spacing:2px; margin-bottom:14px; }
  .avis-client p { font-size:16px; font-style:italic; color:var(--texte); max-width:520px;
        margin:0 auto 16px; line-height:1.6; }
  .avis-client .qui { color:var(--doux); font-size:13px; }
  .contact { background:linear-gradient(135deg, var(--accent), var(--accent-fonce));
             border-radius:22px; padding:40px 28px; text-align:center; color:#fff; }
  .contact h2 { color:#fff; margin:0 0 8px; }
  .contact p { color:rgba(255,255,255,.85); font-size:14.5px; margin:0 0 22px; }
  .contact a.principal { background:#fff; color:var(--accent-fonce); border-color:#fff; box-shadow:none; }
  .contact a.btn:not(.principal) { border-color:rgba(255,255,255,.5); color:#fff; background:transparent; }
  footer { text-align:center; padding:30px 24px 44px; color:var(--doux); font-size:12.5px; }
  footer a { color:var(--accent-fonce); text-decoration:none; font-weight:600; }
</style>
</head>
<body>
<p class="banniere">Aperçu réalisé pour __NOM__ — pas encore un vrai site, un exemple de ce qui est possible en quelques jours.</p>

<nav>
  <span class="marque"><span class="puce">__ICONE__</span>__NOM__</span>
  <div class="cta">
    <a class="btn principal" href="__WA__" target="_blank" rel="noopener">Écrire sur WhatsApp</a>
  </div>
</nav>

<div class="hero">
  __PHOTO_HERO__
  <span class="etiquette">__CATEGORIE__</span>
  <h1>__NOM__</h1>
  <p>__TAGLINE__</p>
  <div class="cta">
    <a class="btn principal" href="__WA__" target="_blank" rel="noopener">Écrire sur WhatsApp</a>
    <a class="btn" href="__CARTE__" target="_blank" rel="noopener">Voir sur la carte</a>
  </div>
  <div class="confiance">
    <span>📍 <b>__ZONE__</b></span>
    <span>📞 <b>__TEL__</b></span>
  </div>
</div>

<main>
  __CORPS__
  __LIEN_MENU__

  <section>
    <h2>Galerie photo</h2>
    <p class="sous-titre">__SOUS_TITRE_GALERIE__</p>
    <div class="galerie">
      __GALERIE__
    </div>
  </section>

  <section>
    <div class="avis-client">
      <div class="etoiles">★★★★★</div>
      <p>« Les avis de vos clients donnent confiance à ceux qui ne vous connaissent pas encore — cet espace est fait pour ça. »</p>
      <div class="qui">Emplacement réservé à un avis client</div>
    </div>
  </section>

  <section>
    <div class="contact">
      <h2>__NOM__</h2>
      <p>__ZONE__ · __TEL__</p>
      <div class="cta">
        <a class="btn principal" href="__WA__" target="_blank" rel="noopener">Écrire sur WhatsApp</a>
        <a class="btn" href="__CARTE__" target="_blank" rel="noopener">Voir sur la carte</a>
      </div>
    </div>
  </section>
</main>

<footer>
  Site d'exemple réalisé par Vision Amah · <a href="https://__SITE__" target="_blank" rel="noopener">__SITE__</a>
</footer>
</body>
</html>
"""


def corps_demo(groupe, sections, elements, wa):
    """Construit le cœur de page — la partie qui change vraiment d'un métier à
    l'autre. Un restaurant vend un menu, un institut des créneaux, une
    boutique un catalogue : recolorer un seul gabarit ne le montrait pas."""
    mise_en_page = MISE_EN_PAGE_GROUPE.get(groupe, "catalogue")
    titre = html.escape(sections[0])
    e = [html.escape(x) for x in elements]

    if mise_en_page == "menu":
        sous_titre = "Vos plats, avec vos vrais prix et descriptions."
        lignes = "".join(
            f'<div class="ligne-menu"><span>{nom}</span><span class="prix">Prix</span></div>'
            for nom in e
        )
        corps = f'<div class="carte menu">{lignes}</div>'

    elif mise_en_page == "rdv":
        sous_titre = "Vos prestations, avec durée et tarif, réservables en direct."
        lignes = "".join(
            f'<div class="service"><div><h4>{nom}</h4><span class="doux">Durée à préciser</span></div>'
            f'<a class="btn mini" href="{wa}" target="_blank" rel="noopener">Réserver</a></div>'
            for nom in e
        )
        corps = f'<div class="liste-services">{lignes}</div>'

    elif mise_en_page == "agenda":
        sous_titre = "Votre programme, toujours à jour pour vos habitués."
        jours = ("VEN", "SAM", "DIM")
        lignes = "".join(
            f'<div class="evenement"><div class="date">{jour}</div>'
            f'<div><h4>{nom}</h4><span class="doux">Horaire à préciser</span></div></div>'
            for jour, nom in zip(jours, e)
        )
        corps = f'<div class="liste-agenda">{lignes}</div>'

    else:  # "catalogue"
        sous_titre = "Vos produits, avec photos et prix à jour."
        cases = "".join(
            f'<div class="produit"><div class="photo-vide"><span>📷</span><span>Photo produit</span></div>'
            f'<h4>{nom}</h4><span class="prix">Prix</span></div>'
            for nom in e
        )
        corps = f'<div class="grille produits">{cases}</div>'

    return f'<section><h2>{titre}</h2><p class="sous-titre">{sous_titre}</p>{corps}</section>'


def photo_hero_html(image, icone):
    if image:
        return (f'<div class="photo-hero reelle" style="background-image:url(\'{html.escape(image)}\')"></div>'
                f'<span class="photo-reelle-etiq">Photo existante (source : OpenStreetMap)</span>')
    return f'<div class="photo-hero">{icone}</div>'


def galerie_html(image):
    """Une vraie photo si OSM en référence une (jamais Google Maps — voir
    `image_osm`), complétée par des emplacements à remplir."""
    case_vide = '<div class="photo-vide"><span>📷</span><span>Votre photo</span></div>'
    if image:
        case_reelle = (f'<div class="photo-vide" style="background-image:url(\'{html.escape(image)}\');'
                        f'background-size:cover;background-position:center;border-style:solid;">'
                        f'<span style="background:rgba(0,0,0,.55);color:#fff;padding:2px 8px;'
                        f'border-radius:99px;font-size:10.5px;">Photo existante</span></div>')
        return case_reelle + case_vide + case_vide
    return case_vide * 3


def lien_menu_html(menu_url):
    if not menu_url:
        return ""
    lien = html.escape(menu_url)
    return (f'<p style="text-align:center;margin:-18px 0 10px;">'
            f'<a class="btn" href="{lien}" target="_blank" rel="noopener">'
            f'Voir le menu déjà en ligne ↗</a></p>')


def rendre_demo(p):
    """Mockup HTML statique, propre au commerce : plus parlant qu'un simple
    message texte pour donner une idée concrète du résultat. À screenshoter
    (ou ouvrir directement sur le téléphone) et joindre au message WhatsApp.
    Utilise une vraie photo et un vrai lien de menu quand OSM en référence,
    plutôt que de toujours partir d'un gabarit entièrement fictif."""
    groupe = p["famille"]
    sections = SECTIONS_GROUPE.get(groupe, SECTIONS_GROUPE["generique"])
    elements = ELEMENTS_GROUPE.get(groupe, ELEMENTS_GROUPE["generique"])
    principal = p["numeros"][0]
    wa = f"https://wa.me/{principal}"
    icone = ICONES_GROUPE.get(groupe, ICONES_GROUPE["generique"])
    theme = THEME_GROUPE.get(groupe, "clair")
    return (GABARIT_DEMO
            .replace("__NOM__", html.escape(p["nom"]))
            .replace("__CATEGORIE__", html.escape(LABELS_GROUPE[groupe]))
            .replace("__ICONE__", icone)
            .replace("__TAGLINE__", html.escape(TAGLINES.get(groupe, TAGLINES["generique"])))
            .replace("__CORPS__", corps_demo(groupe, sections, elements, wa))
            .replace("__LIEN_MENU__", lien_menu_html(p.get("menu_url", "")))
            .replace("__PHOTO_HERO__", photo_hero_html(p.get("image", ""), icone))
            .replace("__GALERIE__", galerie_html(p.get("image", "")))
            .replace("__SOUS_TITRE_GALERIE__",
                     "Une vraie photo trouvée pour vous, à compléter avec les autres."
                     if p.get("image") else
                     "Vos vraies photos remplaceront ces espaces — c'est souvent ce qui convainc le plus.")
            .replace("__WA__", wa)
            .replace("__CARTE__", f'https://www.google.com/maps?q={p["lat"]},{p["lon"]}')
            .replace("__ZONE__", html.escape(p["zone"]))
            .replace("__TEL__", html.escape(" / ".join("+" + n for n in p["numeros"])))
            .replace("__COULEUR__", COULEUR_GROUPE.get(groupe, COULEUR_GROUPE["generique"]))
            .replace("__VARS_THEME__", VARS_THEME[theme])
            .replace("__SITE__", SITE))


def dossier_demos_ecrivable(base="demos"):
    """Un seul test d'écriture pour tout le lot (des centaines de fichiers),
    plutôt qu'un repli par fichier comme `ouvrir_en_ecriture` : on choisit le
    dossier une bonne fois pour toutes, avant de générer quoi que ce soit."""
    for candidat in (base, os.path.join(DOSSIER_APP, base)):
        try:
            os.makedirs(candidat, exist_ok=True)
            test = os.path.join(candidat, ".ecriture_test")
            with open(test, "w"):
                pass
            os.remove(test)
            return candidat
        except OSError:
            continue
    raise PermissionError(
        f"Impossible de créer un dossier pour les démos (ni « {base} », ni « {DOSSIER_APP} »)."
    )


def ecrire_demos(fiches):
    """Génère une page de démo par prospect et note son chemin sur la fiche
    (clé "demo"), pour que le CSV/HTML/PDF puissent y renvoyer."""
    if not fiches:
        return None
    dossier = dossier_demos_ecrivable()
    vus = {}
    for p in fiches:
        base = cle_nom(p["nom"])[:40] or "prospect"
        n = vus.get(base, 0)
        vus[base] = n + 1
        nomfichier = f"{base}.html" if n == 0 else f"{base}-{n + 1}.html"
        chemin = os.path.join(dossier, nomfichier)
        with open(chemin, "w", encoding="utf-8") as f:
            f.write(rendre_demo(p))
        p["demo"] = os.path.abspath(chemin)
    return dossier


# ------------------------------------------------------------------- PDF

def texte_pdf(s):
    """La police de base de fpdf2 (Helvetica) ne connaît que le Latin-1 : les
    emojis des libellés/messages (🍽️, 👋…) la feraient planter. On les retire
    ici plutôt que d'ajouter une police Unicode entière pour ce seul besoin.
    Quelques signes typographiques hors Latin-1 (tiret cadratin, guillemets
    courbes, points de suspension) sont d'abord remplacés par leur équivalent
    ASCII, pour ne pas laisser de double-espace à leur place une fois retirés."""
    s = (s or "")
    # Un précédent copier-coller avait fini par comparer des guillemets
    # droits à eux-mêmes (ne remplaçait donc rien) au lieu des vrais
    # guillemets courbes visés — un nom copié depuis Word (’, “, ”) se
    # retrouvait alors silencieusement amputé de ses apostrophes/guillemets
    # par le .encode("latin-1", "ignore") ci-dessous, plutôt que converti.
    for orig, remplacement in (
        ("—", "-"), ("–", "-"),      # tiret cadratin/demi-cadratin
        ("‘", "'"), ("’", "'"),      # guillemets simples courbes
        ("“", '"'), ("”", '"'),      # guillemets doubles courbes
        ("…", "..."),                     # points de suspension
    ):
        s = s.replace(orig, remplacement)
    return s.encode("latin-1", "ignore").decode("latin-1")


def ecrire_pdf(fiches, chemin=FICHIER_PDF):
    """Sortie principale : contrairement à la page HTML, un PDF s'ouvre de
    façon fiable avec le lecteur par défaut, sans dépendre du navigateur ni
    d'une URL à construire — exactement comme n'importe quel PDF reçu par
    email s'ouvre d'un double-clic."""
    from fpdf import FPDF

    pdf = FPDF(format="A4")
    pdf.set_auto_page_break(auto=True, margin=16)
    pdf.set_title(texte_pdf("Prospection Vision Amah"))
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(15, 15, 15)
    pdf.cell(0, 10, texte_pdf(f"Prospection Vision Amah — {len(fiches)} prospects"), ln=1)
    pdf.set_font("Helvetica", "", 9.5)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6, texte_pdf("Aucun envoi automatique : les liens WhatsApp ouvrent une conversation à valider soi-même."), ln=1)
    pdf.ln(4)

    par_groupe = {}
    for f in fiches:
        par_groupe.setdefault(f["famille"], []).append(f)

    for g in ORDRE_GROUPES:
        items = par_groupe.get(g)
        if not items:
            continue

        pdf.set_font("Helvetica", "B", 13)
        pdf.set_text_color(15, 15, 15)
        # .lstrip() : retire l'espace laissé par l'emoji du libellé, supprimé par texte_pdf.
        pdf.cell(0, 9, texte_pdf(f"{LABELS_GROUPE[g]} ({len(items)})").lstrip(), ln=1)
        pdf.set_draw_color(200, 200, 200)
        pdf.line(pdf.get_x(), pdf.get_y(), pdf.get_x() + 180, pdf.get_y())
        pdf.ln(3)

        for p in items:
            principal = p["numeros"][0]
            wa = f"https://wa.me/{principal}"
            carte = f'https://www.google.com/maps?q={p["lat"]},{p["lon"]}'
            tel = " / ".join("+" + n for n in p["numeros"])

            pdf.set_font("Helvetica", "B", 11)
            pdf.set_text_color(10, 10, 10)
            etiquette = " (prioritaire)" if score(p) else ""
            pdf.cell(0, 6, texte_pdf(p["nom"] + etiquette), ln=1)

            pdf.set_font("Helvetica", "", 9.5)
            pdf.set_text_color(90, 90, 90)
            pdf.cell(0, 5, texte_pdf(f'{p["zone"]} - {p["type_osm"]} - {tel}'), ln=1)

            pdf.set_font("Helvetica", "U", 9.5)
            pdf.set_text_color(37, 99, 235)
            pdf.cell(40, 5.5, texte_pdf("WhatsApp"), link=wa)
            if p.get("demo"):
                pdf.cell(40, 5.5, texte_pdf("Voir sur la carte"), link=carte)
                pdf.cell(0, 5.5, texte_pdf("Voir la démo"), link=pathlib.Path(p["demo"]).as_uri(), ln=1)
            else:
                pdf.cell(0, 5.5, texte_pdf("Voir sur la carte"), link=carte, ln=1)

            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(60, 60, 60)
            pdf.multi_cell(0, 4.6, texte_pdf(p["message"]))
            pdf.ln(4)

    # Comme pour le CSV/HTML : si le dossier courant refuse l'écriture (Contrôle
    # d'accès aux dossiers, ou fichier ouvert dans un lecteur PDF), on retombe
    # sur un nom horodaté au même endroit, puis sur le dossier applicatif.
    for cible in (chemin, chemin_de_repli(chemin), os.path.join(DOSSIER_APP, os.path.basename(chemin))):
        try:
            pdf.output(cible)
            if cible != chemin:
                print(f"[!] Impossible d'écrire dans « {chemin} » — écrit dans « {cible} » à la place.")
            return cible
        except PermissionError:
            continue
    raise PermissionError(f"Impossible d'écrire {chemin} : dossier bloqué en écriture, y compris {DOSSIER_APP}.")


# ----------------------------------------------------------------------- main

def main(ville="Conakry"):
    ajoutees = fusionner_journal_envois()
    if ajoutees:
        print(f"Journal des envois : {ajoutees} entrée(s) ajoutée(s) ({os.path.abspath(FICHIER_JOURNAL_XLSX)})\n")
    elif ajoutees == 0:
        print(f"Journal des envois : export « {FICHIER_JOURNAL_EXPORT} » trouvé, rien de nouveau à ajouter.\n")

    elements = interroger_overpass(ville)
    print(f"{len(elements)} résultats bruts")

    fiches, rejets = extraire(elements)
    print(f"  − {rejets['sans_nom']} sans nom (impossible à personnaliser)")
    print(f"  − {rejets['sans_numero_valide']} sans numéro exploitable")

    fiches, doublons = dedupliquer(fiches)
    for nom, raison in doublons:
        print(f"  − doublon écarté : {nom} ({raison})")

    connues = charger_fiches_connues()
    nouvelles = [f for f in fiches if f["id_osm"] not in connues]
    print(f"  − {len(fiches) - len(nouvelles)} déjà connus d'une exécution précédente")

    for f in nouvelles:
        f["zone"] = zone_la_plus_proche(f["lat"], f["lon"])
        connues[f["id_osm"]] = f

    # On régénère TOUJOURS à partir de l'ensemble connu au complet (anciens +
    # nouveaux), jamais des seuls nouveaux : sinon prospects.html perdrait
    # chaque semaine — à chaque lancement de la tâche planifiée — tous les
    # prospects des exécutions précédentes, avec leur statut/relance/stats.
    # Le message est reconstruit à chaque fois pour tout le monde, pour que
    # les anciens profitent aussi des derniers gabarits de message.
    toutes = list(connues.values())
    for f in toutes:
        f["message"] = MESSAGES[f["famille"]].format(
            nom=f["nom"], site=SITE, signature=SIGNATURE
        )

    toutes.sort(key=lambda f: (ORDRE_GROUPES.index(f["famille"]), f["zone"], -score(f), f["nom"]))

    dossier_demos = ecrire_demos(toutes)
    chemin_csv = ecrire_csv(toutes)
    chemin_html = ecrire_html(toutes)
    chemin_pdf = ecrire_pdf(toutes)
    sauvegarder_fiches_connues(connues)

    print(f"\n=== {len(toutes)} prospects au total ({len(nouvelles)} nouveaux) ===")
    par_groupe = {}
    for f in toutes:
        par_groupe.setdefault(f["famille"], []).append(f)
    for g in ORDRE_GROUPES:
        if g in par_groupe:
            lst = par_groupe[g]
            print(f"  {LABELS_GROUPE[g]:<22} {len(lst):>3}  ({sum(1 for f in lst if score(f))} prioritaires)")

    print(f"\nPDF (ouvert automatiquement) : {os.path.abspath(chemin_pdf)}")
    print(f"CSV (classement/sauvegarde)  : {os.path.abspath(chemin_csv)}")
    print(f"Page interactive (facultatif): {os.path.abspath(chemin_html)}")
    if dossier_demos:
        print(f"Pages de démo (une par prospect) : {os.path.abspath(dossier_demos)}")

    # Le PDF est le format ouvert automatiquement : contrairement à la page
    # HTML, il ne dépend pas du navigateur ni d'une URL à construire. On garde
    # `os.startfile()` — le mécanisme natif de Windows, celui de l'Explorateur
    # au double-clic — plutôt que `webbrowser.open()`, plus capricieux.
    chemin_absolu = os.path.abspath(chemin_pdf)
    ouvert = False
    if hasattr(os, "startfile"):
        try:
            os.startfile(chemin_absolu)
            ouvert = True
        except OSError as e:
            print(f"[!] os.startfile a échoué ({e}).")
    if not ouvert:
        try:
            webbrowser.open(pathlib.Path(chemin_absolu).as_uri())
        except Exception as e:
            print(f"[!] Ouverture automatique impossible ({e}).")
    print(f"Si rien ne s'ouvre, double-clique ce fichier à la main : {chemin_absolu}")


if __name__ == "__main__":
    # Un double-clic sur ce fichier ouvre une fenêtre qui se referme SEULE dès
    # que le programme s'arrête — succès ou erreur, aucune différence. Sans ce
    # bloc, une erreur s'affiche et disparaît en une fraction de seconde,
    # impossible à lire. On l'affiche donc en entier et on attend une touche
    # avant de fermer, dans tous les cas.
    try:
        main(sys.argv[1] if len(sys.argv) > 1 else "Conakry")
        print("\n=== Terminé avec succès ===")
    except requests.exceptions.RequestException as e:
        print(f"\n[!] Impossible de joindre OpenStreetMap : {e}")
        print("Vérifie ta connexion internet, puis relance le programme.")
    except Exception:
        import traceback
        print("\n[!] Le programme s'est arrêté à cause d'une erreur :\n")
        traceback.print_exc()
        print("\nEnvoie ce message d'erreur pour obtenir de l'aide.")
    finally:
        # Une tâche planifiée (ou tout lancement sans console attachée) n'a
        # pas d'entrée interactive — attendre une touche bloquerait alors
        # indéfiniment le programme. isatty() filtre la plupart des cas ;
        # EOFError couvre les terminaux où isatty() répond True à tort mais
        # où aucune touche ne peut de toute façon être tapée.
        if sys.stdin.isatty():
            try:
                input("\nAppuie sur Entrée pour fermer cette fenêtre...")
            except EOFError:
                pass
